"""
SLGS Skill Matcher — Hybrid (Rule-based + Claude API)
Reads Excel data, matches available people to open roles by skill similarity.
Outputs JSON results consumable by the dashboard HTML generator.

Usage:
    python skill_matcher.py

Requires:
    pip install openpyxl anthropic
"""

import os
import json
import re
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).parent
OUTPUT_FILE = DATA_DIR / "skill_matches.json"

# Claude API — only used for ambiguous matches (hybrid mode)
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
USE_AI_MATCHING = bool(ANTHROPIC_API_KEY)
AI_MODEL = "claude-sonnet-4-6-20250514"

# Minimum keyword overlap score to consider a match (0-1)
KEYWORD_THRESHOLD = 0.3
# Maximum people to return per role
MAX_MATCHES_PER_ROLE = 5

# ---------------------------------------------------------------------------
# DATA LOADING
# ---------------------------------------------------------------------------


def load_supply_pool():
    """Load available people (bench/pipeline/active) with their skills."""
    people = []

    # Source 1: India CD & US Tech — has detailed people data
    wb_path = DATA_DIR / "India CD & US Tech.xlsx"
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        ws = wb["Data"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]

        col = {h: i for i, h in enumerate(headers) if h}

        for row in rows[1:]:
            name = row[col.get("Employee Name", 14)]
            status = row[col.get("Status", 26)]
            category = row[col.get("Category", 24)] if len(row) > 24 else None
            primary = row[col.get("Primary Skills", 19)]
            secondary = row[col.get("Secondary Skills", 20)]
            team = row[col.get("Team Name", 16)]
            function = row[col.get("Function", 1)]
            new_role_skill = row[col.get("Skills for the new role", 36)] if len(row) > 36 else None
            emp_type = row[col.get("Old Employee Type", 39)] if len(row) > 39 else None

            if not name:
                continue

            # Only include active/available people
            is_available = (
                status == "Active"
                or category in ("In Pipeline Capacity Creation", "Moved to a new role", "Active")
            )
            if not is_available:
                continue

            people.append({
                "name": str(name).strip(),
                "primary_skills": str(primary).strip() if primary else "",
                "secondary_skills": str(secondary).strip() if secondary else "",
                "team": str(team).strip() if team else "",
                "function": str(function).strip() if function else "",
                "new_role_skill": str(new_role_skill).strip() if new_role_skill else "",
                "employee_type": str(emp_type).strip() if emp_type else "",
                "source": "India CD & US Tech",
                "status": str(status),
                "category": str(category) if category else "",
            })
        wb.close()

    # Source 2: India Ops — similar structure
    wb_path = DATA_DIR / "India Ops.xlsx"
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        col = {h: i for i, h in enumerate(headers) if h}

        for row in rows[1:]:
            # India Ops may have employee names in similar columns
            name_idx = col.get("Employee Name")
            if name_idx and len(row) > name_idx and row[name_idx]:
                people.append({
                    "name": str(row[name_idx]).strip(),
                    "primary_skills": str(row[col.get("Primary Skills", -1)]).strip() if col.get("Primary Skills") and len(row) > col["Primary Skills"] else "",
                    "secondary_skills": "",
                    "team": "",
                    "function": "Operations",
                    "new_role_skill": "",
                    "employee_type": "",
                    "source": "India Ops",
                    "status": "Active",
                    "category": "",
                })
        wb.close()

    return people


def load_demand_roles():
    """Load open roles/positions that need filling."""
    roles = []

    # Source 1: India TA Open Positions
    wb_path = DATA_DIR / "India TA Open Positions Data - May 8.xlsx"
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        col = {h: i for i, h in enumerate(headers) if h}

        for row in rows[1:]:
            status = row[col.get("Status", 17)] if col.get("Status") else None
            if status != "Open":
                continue
            skillset = row[col.get("Skillset", 4)]
            skill_type = row[col.get("Skill Type", 3)]
            role_title = row[col.get("Designation", 12)] if col.get("Designation") else ""
            dept = row[col.get("Department", 10)] if col.get("Department") else ""
            location = row[col.get("Location", 5)] if col.get("Location") else ""

            if not skillset:
                continue

            roles.append({
                "title": str(role_title).strip() if role_title else str(skillset).strip(),
                "required_skills": str(skillset).strip(),
                "skill_type": str(skill_type).strip() if skill_type else "Vanilla",
                "department": str(dept).strip() if dept else "",
                "location": str(location).strip() if location else "",
                "source": "India TA",
            })
        wb.close()

    # Source 2: PH Open Demand
    wb_path = DATA_DIR / "SLGS PH Open Demand_260511.xlsx"
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        ws = wb["Open DRFs (For HR)"]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row (row with 'Demand Type' in first col)
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] == "Demand Type":
                header_idx = i
                break

        if header_idx:
            headers = rows[header_idx]
            col = {h: i for i, h in enumerate(headers) if h}

            for row in rows[header_idx + 1:]:
                if not row or not row[0] or "Total" in str(row[0]):
                    continue
                role_title = row[col.get("Role of Requested Demand", 4)] if col.get("Role of Requested Demand") else ""
                primary_skill = row[col.get("Required Skill\r\n(Primary)", 7)] if len(row) > 7 else ""
                secondary_skill = row[col.get("Required Skill\r\n(Secondary)", 8)] if len(row) > 8 else ""
                market = row[col.get("Market Assignment", 2)] if col.get("Market Assignment") else ""

                if not primary_skill or primary_skill == "(blank)":
                    continue

                combined_skills = str(primary_skill)
                if secondary_skill and secondary_skill != "(blank)":
                    combined_skills += ", " + str(secondary_skill)

                roles.append({
                    "title": str(role_title).strip() if role_title else "Open Role",
                    "required_skills": combined_skills.strip(),
                    "skill_type": "Niche",
                    "department": "",
                    "location": str(market).strip() if market else "PH",
                    "source": "PH DRF",
                })
        wb.close()

    # Source 3: India Contractor Positions
    wb_path = DATA_DIR / "India Contractor Positions - 13 May.xlsx"
    if wb_path.exists():
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        ws = wb["Open & Offered Roles"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 2:
            headers = rows[1]
            col = {h: i for i, h in enumerate(headers) if h}

            for row in rows[2:]:
                if not row or not row[0]:
                    continue
                role_title = row[col.get("Role", 3)] if col.get("Role") else ""
                primary_skill = row[col.get("Primary Skill", 8)] if col.get("Primary Skill") else ""
                secondary_skill = row[col.get("Secondary Skill", 9)] if col.get("Secondary Skill") else ""

                if not primary_skill:
                    continue

                combined = str(primary_skill)
                if secondary_skill:
                    combined += ", " + str(secondary_skill)

                roles.append({
                    "title": str(role_title).strip() if role_title else "Contractor Role",
                    "required_skills": combined.strip(),
                    "skill_type": "Niche",
                    "department": "Technology",
                    "location": "Gurugram",
                    "source": "India Contractor",
                })
        wb.close()

    return roles


# ---------------------------------------------------------------------------
# KEYWORD MATCHING (Rule-based — fast, no API)
# ---------------------------------------------------------------------------


def tokenize_skills(skill_string):
    """Break skill string into normalized tokens for matching."""
    if not skill_string:
        return set()
    # Normalize separators
    text = skill_string.lower()
    text = re.sub(r"[/,;|•\n\r]+", " ", text)
    text = re.sub(r"[^a-z0-9\s.#+]", " ", text)
    tokens = set(text.split())
    # Remove noise words
    noise = {"and", "or", "the", "in", "of", "for", "with", "a", "an", "to", "is", "on", "at", "by"}
    tokens -= noise
    # Also keep multi-word phrases
    phrases = re.findall(r"[a-z][a-z ]{3,}", text)
    for p in phrases:
        tokens.add(p.strip())
    return tokens


def keyword_match_score(person_skills, role_skills):
    """Score 0-1 based on token overlap between person's skills and role requirements."""
    person_tokens = tokenize_skills(person_skills)
    role_tokens = tokenize_skills(role_skills)

    if not role_tokens or not person_tokens:
        return 0.0

    overlap = person_tokens & role_tokens
    # Jaccard-like but weighted toward role coverage
    role_coverage = len(overlap) / len(role_tokens) if role_tokens else 0
    return round(role_coverage, 3)


def match_rule_based(people, roles):
    """First pass: keyword matching. Returns matches above threshold."""
    results = []

    for role in roles:
        role_matches = []
        for person in people:
            # Combine person's skills
            all_person_skills = " ".join([
                person["primary_skills"],
                person["secondary_skills"],
                person["new_role_skill"],
            ])

            score = keyword_match_score(all_person_skills, role["required_skills"])

            if score >= KEYWORD_THRESHOLD:
                role_matches.append({
                    "name": person["name"],
                    "score": score,
                    "confidence": "high" if score > 0.7 else "medium" if score > 0.5 else "low",
                    "basis": f"Keyword overlap: {score:.0%}",
                    "primary_skills": person["primary_skills"],
                    "source": person["source"],
                    "category": person["category"],
                    "match_type": "rule_based",
                })

        # Sort by score descending, take top N
        role_matches.sort(key=lambda x: x["score"], reverse=True)
        results.append({
            "role": role,
            "matches": role_matches[:MAX_MATCHES_PER_ROLE],
            "unmatched": len(role_matches) == 0,
        })

    return results


# ---------------------------------------------------------------------------
# AI MATCHING (Claude API — for ambiguous/unmatched roles)
# ---------------------------------------------------------------------------


def match_with_claude(unmatched_roles, people):
    """Second pass: use Claude to semantically match remaining unmatched roles."""
    if not USE_AI_MATCHING:
        return []

    try:
        from anthropic import Anthropic
    except ImportError:
        print("WARNING: anthropic package not installed. Skipping AI matching.")
        return []

    client = Anthropic(api_key=ANTHROPIC_API_KEY)

    # Build people summary for context
    people_summary = "\n".join([
        f"- {p['name']}: {p['primary_skills']} | {p['secondary_skills']} | Team: {p['team']}"
        for p in people if p["primary_skills"]
    ])

    ai_results = []

    # Batch unmatched roles (max 10 per call to stay efficient)
    for batch_start in range(0, len(unmatched_roles), 10):
        batch = unmatched_roles[batch_start:batch_start + 10]

        roles_text = "\n".join([
            f"Role {i+1}: {r['role']['title']} — Skills needed: {r['role']['required_skills']}"
            for i, r in enumerate(batch)
        ])

        prompt = f"""You are an internal HR skill-matching assistant for Sun Life (SLGS).

Given these OPEN ROLES that need filling:
{roles_text}

And this AVAILABLE TALENT POOL:
{people_summary}

For each role, identify the top 1-3 best-matching people from the talent pool.
Consider:
- Direct skill matches (exact keywords)
- Adjacent/transferable skills (e.g., "AWS Glue" person can do "ETL Engineering")
- Team/function relevance

Return JSON array:
[
  {{
    "role_index": 1,
    "matches": [
      {{"name": "Person Name", "score": 0.75, "reasoning": "brief explanation"}}
    ]
  }}
]

Only return the JSON, no other text."""

        try:
            response = client.messages.create(
                model=AI_MODEL,
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}]
            )
            content = response.content[0].text

            # Parse JSON from response
            json_match = re.search(r"\[.*\]", content, re.DOTALL)
            if json_match:
                ai_matches = json.loads(json_match.group())
                for match_result in ai_matches:
                    idx = match_result.get("role_index", 1) - 1
                    if 0 <= idx < len(batch):
                        for m in match_result.get("matches", []):
                            batch[idx].setdefault("matches", []).append({
                                "name": m["name"],
                                "score": m.get("score", 0.6),
                                "confidence": "medium",
                                "basis": m.get("reasoning", "AI semantic match"),
                                "primary_skills": "",
                                "source": "AI Match",
                                "category": "",
                                "match_type": "ai_semantic",
                            })
                        batch[idx]["unmatched"] = False
        except Exception as e:
            print(f"WARNING: Claude API call failed: {e}")

        ai_results.extend(batch)

    return ai_results


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------


def run():
    print("Loading supply pool (available people)...")
    people = load_supply_pool()
    print(f"  Found {len(people)} available people")

    print("Loading demand roles (open positions)...")
    roles = load_demand_roles()
    print(f"  Found {len(roles)} open roles")

    print("\nPass 1: Rule-based keyword matching...")
    results = match_rule_based(people, roles)
    matched_count = sum(1 for r in results if r["matches"])
    unmatched = [r for r in results if r["unmatched"]]
    print(f"  Matched: {matched_count} roles")
    print(f"  Unmatched: {len(unmatched)} roles")

    if USE_AI_MATCHING and unmatched:
        print(f"\nPass 2: AI semantic matching for {len(unmatched)} unmatched roles...")
        ai_results = match_with_claude(unmatched, people)
        # Merge AI results back
        for ai_r in ai_results:
            for r in results:
                if r["role"] == ai_r["role"]:
                    r["matches"] = ai_r.get("matches", [])
                    r["unmatched"] = ai_r.get("unmatched", True)
                    break
        newly_matched = sum(1 for r in ai_results if not r.get("unmatched"))
        print(f"  AI matched: {newly_matched} additional roles")
    elif not USE_AI_MATCHING and unmatched:
        print("\n  (Set ANTHROPIC_API_KEY env var to enable AI matching for unmatched roles)")

    # Build output
    output = {
        "generated_at": "2026-05-22",
        "summary": {
            "total_people": len(people),
            "total_roles": len(roles),
            "roles_matched": sum(1 for r in results if r["matches"]),
            "roles_unmatched": sum(1 for r in results if not r["matches"]),
            "matching_mode": "hybrid" if USE_AI_MATCHING else "rule_based",
        },
        "matches": [
            {
                "role_title": r["role"]["title"],
                "required_skills": r["role"]["required_skills"],
                "source": r["role"]["source"],
                "location": r["role"]["location"],
                "candidates": r["matches"],
            }
            for r in results
            if r["matches"]  # Only include roles that have matches
        ],
        "unmatched_roles": [
            {
                "role_title": r["role"]["title"],
                "required_skills": r["role"]["required_skills"],
                "source": r["role"]["source"],
            }
            for r in results
            if not r["matches"]
        ],
    }

    # Write output
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nResults written to: {OUTPUT_FILE}")
    print(f"\nTop matches preview:")
    for match in output["matches"][:5]:
        print(f"\n  Role: {match['role_title']} ({match['source']})")
        print(f"  Needs: {match['required_skills'][:60]}")
        for c in match["candidates"][:3]:
            print(f"    → {c['name']} | Score: {c['score']:.0%} | {c['basis']}")


if __name__ == "__main__":
    run()
